from openpyxl import Workbook, load_workbook
import timestamp, json

strr=['document-uri','referrer','violated-directive','effective-directive','original-policy','disposition','blocked-uri','line-number','column-number','source-file','status-code','script-sample']



def deserialize(filePath):
    try:
        text_file = open(filePath)
        data = text_file.readlines()
        text_file.close()
        data=str(data[0])
        data=data.split('{"csp-report":',1)[1]
        data=data[:-1]
        data=json.loads(data)
        return data
    except:
        print("An exception occurred while opening the file")

def wrtDataToXl(r,c,finalStr):
    try:
        sheet.cell(row=r, column=c).value = finalStr  
    except:
        print("An exception occurred while writing data to the excel")

def sort(strr,data,r):
    try:
        c=1
        for i in strr:
            # print(data.get(i))
            wrtDataToXl(r,c,str(data.get(i)))
            c=c+1
    except:
        print("an error occured during writing data to excel")

    
def genReport(filePath,tmpdir):
    pathToCsvFile=tmpdir+'/SpreadSheet_'+timestamp.timestamp()+'.xlsx'
    wb = Workbook()
    wb.save(pathToCsvFile)
    wb = load_workbook(pathToCsvFile)
    global sheet
    sheet = wb.active
    count=2
    c=1
    for i in strr:
        wrtDataToXl(1,c,i)
        c=c+1
        
    for i in filePath:
        data=(deserialize(i))
        sort(strr,data,count)
        count+=1
    wb.save(pathToCsvFile)
    return str(pathToCsvFile)